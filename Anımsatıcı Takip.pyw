import time
import datetime
from openpyxl import load_workbook
from plyer import notification

while True:

    saat = 0
    dakika = 0

    try:
        listeA = []
        listeB = []
        dt = datetime.datetime.now()
        hour = dt.hour
        minute = dt.minute
        wb = load_workbook(r"Anımsatıcılar.xlsx")
        ws = wb.active
        for i in range(1,10):
            i = str(i)
            if ((ws["A"+i] != "") and (ws["B"+i] != "")):
                listeA.append(ws["A"+i].value)
                listeB.append(ws["B"+i].value)
        for saatim in listeA:
            for dakkam in listeB:
                if ((saatim != None) and (dakkam != None)):
                    if ((hour == saatim) and (minute == dakkam)):
                        notification.notify(title="Joseph-Bot",message="Saat ayarladığınız zamana geldi.", app_name = "Joseph-Bot")
                        old_minute = minute
                        while True:
                            hour = dt.hour
                            minute = dt.minute
                            if old_minute != minute:
                                break

                    
                    if ((hour == 12) and (minute == 0)):
                        notification.notify(title="Joseph-Bot",message="Saat 12:00 oldu.", app_name = "Joseph-Bot")
                        time.sleep(60)

        time.sleep(0.1)
    except:
        notification.notify(title="Joseph-Bot",message="Bir hata meydana geldi...", app_name = "Joseph-Bot")
        exit()