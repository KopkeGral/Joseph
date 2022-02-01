from tkinter import *
from tkinter import messagebox
import requests
from bs4 import BeautifulSoup
import datetime as dt
import random
import webbrowser as wb
import os
import time
from openpyxl import *
from googletrans import Translator
translator = Translator()

workbook = load_workbook("Liste.xlsx")
sheet = workbook.active

diller = ['afrikaans', 'albanian', 'amharic', 'arabic', 'armenian', 'azerbaijani', 'basque', 'belarusian', 'bengali', 'bosnian', 'bulgarian', 'catalan', 'cebuano', 'chichewa', 'chinese (simplified)', 'chinese (traditional)', 'corsican', 'croatian', 'czech', 'danish', 'dutch', 'english', 'esperanto', 'estonian', 'filipino', 'finnish', 'french', 'frisian', 'galician', 'georgian', 'german', 'greek', 'gujarati', 'haitian creole', 'hausa', 'hawaiian', 'hebrew', 'hebrew', 'hindi', 'hmong', 'hungarian', 'icelandic', 'igbo', 'indonesian', 'irish', 'italian', 'japanese', 'javanese', 
'kannada', 'kazakh', 'khmer', 'korean', 'kurdish (kurmanji)', 'kyrgyz', 'lao', 'latin', 'latvian', 'lithuanian', 'luxembourgish', 'macedonian', 'malagasy', 'malay', 'malayalam', 'maltese', 'maori', 'marathi', 'mongolian', 'myanmar (burmese)', 'nepali', 'norwegian', 'odia', 'pashto', 'persian', 'polish', 'portuguese', 'punjabi', 'romanian', 'russian', 'samoan', 'scots gaelic', 'serbian', 'sesotho', 'shona', 'sindhi', 'sinhala', 'slovak', 'slovenian', 'somali', 'spanish', 'sundanese', 'swahili', 'swedish', 'tajik', 'tamil', 'telugu', 'thai', 'turkish', 'ukrainian', 'urdu', 'uyghur', 'uzbek', 'vietnamese', 'welsh', 'xhosa', 'yiddish', 'yoruba', 'zulu']

selam_joseph = [

    "Joseph selamlamanın onurlu bir savaşçının kılıcı olduğuna inanıyor, sana da merhaba!",
    "Joseph bu kibar insanı selamlıyor!",
    "Joseph soruyor: BU KİBARLIK DA NEREDEN GELİYOR?"

]

espriler = [

    "Geçen gün bir taksi çevirdim, hala dönüyor. :)",
    "Yıkanmakta olan balığa ne denir?\nWashington. :)",
    "Bebeğe patik giydirmeye çalışmışlar ama giydirememişler, neden?\nÇünkü bebek antipatikmiş. :)",
    "Hangi örtüyü masaya seremeyiz?\nBitki örtüsü. :)",
    "İshal olmuş böceğe ne denir?\nCırcır böceği. :)",
    "-Nuri ölünce Çin'e gömsünler.\n-Neden?\n-Nuriçinde yatsın diye. :)",
    "Baraj dendi mi, akan sular durur. :)",
    "Hava korsanı uçağı kaçıracaktı ama kaçıramadı çünkü uçağı kaçırdı. :)",
    "-En yeni şehrimiz hangisidir?\n-Hangisidir?\n-Nevşehir. :)",
    "Röntgen filmi çektirdik, yakında sinemalarda! :)",
    "Seven unutmaz, eight unutur. :)",
    "Köstebekler, ben beklemem aga. :)",
    "Baykuşlar vedalaşırken ne der?\n'BAY BAY BAYKUŞ' :)",
    "-Gülen ördeğe ne denir?\n-Ne denir\n-Kıkır-duck :)"

]

def selamla():
    selam = random.choice(selam_joseph)
    messagebox.showinfo("Selam",selam)

def google_ara(arama):
    url = "https://www.google.com/search?q="+arama
    wb.open(url)
    messagebox.showinfo("Google Arama", message="Google arama başarılı.")

def youtube_ara(arama):
    url = "https://www.youtube.com/search?q="+arama
    wb.open(url)
    messagebox.showinfo("Youtube Arama", message="Youtube arama başarılı.")

def wiki_arat():
    wiki_entry_get = wiki_entry.get()
    if wiki_entry_get != "":
        wikipedia(wiki_entry_get)
        
    else:
        messagebox.showerror("Hata","Lütfen wikipedia.org üzerinde aratılmak üzere\nbir değer giriniz.")
        
    wiki_toplevel.destroy()
        
def start_wikipedia():
    global wiki_entry
    global wiki_toplevel
    wiki_toplevel = Toplevel()
    wiki_toplevel.geometry("350x250")
    wiki_toplevel.title("Joseph/Wikipedia")
    wiki_toplevel.minsize(350,250)
    wiki_entry = Entry(wiki_toplevel, bg="gray", border=2)
    wiki_entry.place(relx=0.5, rely=0.1, anchor=CENTER, relheight=0.08, relwidth=0.7)
    wiki_button = Button(wiki_toplevel, text="Arat", command=wiki_arat, bg="#00f000")
    wiki_button.place(relx=0.5, rely=0.80, anchor=CENTER, relheight=0.15, relwidth=0.2)

def work(konu):
    r = requests.get("https://www.google.com/search?q="+konu)
    soup = BeautifulSoup(r.content,features="lxml")

    wbfirst = "https://www.google.com/"

    list = []

    linkler = soup.find_all("a")
    for link in linkler:
        s = link.get("href")
        s = str(s)
        list.append(s)

    for i in list:
        if (i.startswith("/url")):
            break
    messagebox.showinfo("Joseph", message="Url açılıyor: {}".format(wbfirst+i))
    wb.open(wbfirst+i)

def wikipedia(konu):
    global sonlink
    r = requests.get("https://tr.wikipedia.org/w/index.php?search={}&title=Özel:Ara&profile=advanced&fulltext=1&ns0=1".format(konu))
    soup = BeautifulSoup(r.content,features="lxml")

    list = []

    link = soup.find_all("div", attrs={"class":"mw-search-result-heading"})
    link = str(link)
    link = link.split("href=")
    for i in link:
        if "/wiki/" in i:
            break
    
    point = 0
    pointy = '"'
    for j in i:
        if point == 2:
            break
        if j == pointy:
            point = point+1
        list.append(j)
        
    sonlink = ""
    list.pop(0)
    list.pop((len(list)-1))
    for v in list:
        sonlink = sonlink+v
    if sonlink == "":
        messagebox.showerror("Hata","Wikipedia sonuçları arasında {} şeklinde bir sonuç bulunamadı,\nlütfen daha sonra tekrar deneyin veya internetinizi kontrol edin.".format(konu))
    else:
        wb.open("https://tr.wikipedia.org"+sonlink)

def games():
    global games_toplevel
    games_toplevel = Toplevel(bg="#d9d9d9")
    games_toplevel.geometry("350x250")
    games_toplevel.title("Joseph/Games")
    games_toplevel.minsize(350,250)
    Button(games_toplevel,text="Zar at", command=zar_at).place(anchor=CENTER, relx=0.5, rely=0.1, relwidth=0.7, relheight=0.09)

def zar_at():
    zar_first_int = 1
    zar_second_int = 6
    zar_result = random.randint(zar_first_int, zar_second_int)
    messagebox.showinfo("Joseph", message="Zar sonucu: {}".format(zar_result))
    games_toplevel.destroy()




def shutdown():
    if messagebox.askokcancel("Kapat","Bilgisayarı kapatmak istediğinize emin misiniz?"):
        messagebox.showwarning("Kapat","Bilgisayar kapatılıyor...")
        os.system('shutdown -s')

def restart():
    if messagebox.askokcancel("Yeniden başlat","Bilgisayarı yeniden başlatmak istediğinize emin misiniz?"):
        messagebox.showwarning("Yeniden başlat","Bilgisayar yeniden başlatılıyor...")
        os.system('shutdown -r')

def shutdown_and_countdown():
    saniye = 0
    def atmış():
        saniye = 60
        messagebox.showinfo("Zamanlayıcılı Kapatma","{} saniye içerisinde kapatılacaktır...".format(saniye))
        os.system('shutdown -s -t {}'.format(saniye))
        shutdown_toplevel.destroy()
    def otuz():
        saniye = 30
        messagebox.showinfo("Zamanlayıcılı Kapatma","{} saniye içerisinde kapatılacaktır...".format(saniye))
        os.system('shutdown -s -t {}'.format(saniye))
        shutdown_toplevel.destroy()
    def onbeş():
        saniye = 15
        messagebox.showinfo("Zamanlayıcılı Kapatma","{} saniye içerisinde kapatılacaktır...".format(saniye))
        os.system('shutdown -s -t {}'.format(saniye))
        shutdown_toplevel.destroy()
    def beş():
        saniye = 5
        messagebox.showinfo("Zamanlayıcılı Kapatma","{} saniye içerisinde kapatılacaktır...".format(saniye))
        os.system('shutdown -s -t {}'.format(saniye))
        shutdown_toplevel.destroy()
    def bir():
        saniye = 1
        messagebox.showinfo("Zamanlayıcılı Kapatma","{} saniye içerisinde kapatılacaktır...".format(saniye))
        os.system('shutdown -s -t {}'.format(saniye))
        shutdown_toplevel.destroy()

    shutdown_toplevel = Toplevel(bg="#d9d9d9")
    shutdown_toplevel.geometry("350x250")
    shutdown_toplevel.title("Joseph/Bilgisayar İşlemleri/Zamanlayıcılı Kapatma")
    shutdown_toplevel.minsize(350,250)
    
    Button(shutdown_toplevel,text="60 saniye", command=atmış).place(anchor=CENTER, relx=0.5, rely=0.1, relwidth=0.7, relheight=0.09)
    Button(shutdown_toplevel,text="30 saniye", command=otuz).place(anchor=CENTER, relx=0.5, rely=0.2, relwidth=0.7, relheight=0.09)
    Button(shutdown_toplevel,text="15 saniye", command=onbeş).place(anchor=CENTER, relx=0.5, rely=0.3, relwidth=0.7, relheight=0.09)
    Button(shutdown_toplevel,text="5 saniye", command=beş).place(anchor=CENTER, relx=0.5, rely=0.4, relwidth=0.7, relheight=0.09)
    Button(shutdown_toplevel,text="1 saniye", command=bir).place(anchor=CENTER, relx=0.5, rely=0.5, relwidth=0.7, relheight=0.09)
        
def computer_os():
    computer_os_toplevel = Toplevel(bg="#d9d9d9")
    computer_os_toplevel.geometry("350x250")
    computer_os_toplevel.title("Joseph/Bilgisayar İşlemleri")
    computer_os_toplevel.minsize(350,250)
    Button(computer_os_toplevel,text="Bilgisayarı Kapat", command=shutdown).place(anchor=CENTER, relx=0.5, rely=0.1, relwidth=0.7, relheight=0.09)
    Button(computer_os_toplevel,text="Yeniden Başlat", command=restart).place(anchor=CENTER, relx=0.5, rely=0.34, relwidth=0.7, relheight=0.09)
    Button(computer_os_toplevel,text="Zamanlayıcılı Kapatma", command=shutdown_and_countdown).place(anchor=CENTER, relx=0.5, rely=0.22, relwidth=0.7, relheight=0.09)

def ceviri():
    global lang_opsiyon
    global english_textbox
    global turkce_textbox
    translate_toplevel = Toplevel()
    translate_toplevel.title("Translate")
    translate_toplevel.geometry("1000x300")
    translate_toplevel.minsize(1000,300)
    translate_toplevel.maxsize(2000,600)

    english_textbox= Text(translate_toplevel)
    english_textbox.place(relx=0.05,rely=0.1, relwidth=0.4, relheight=0.27)
    turkce_textbox= Text(translate_toplevel)
    turkce_textbox.place(relx=0.55,rely=0.1, relwidth=0.4, relheight=0.27)

    btn1 = Button(translate_toplevel, text="Çevir",font="Times 13 bold",command=encevir)
    btn1.place(relx=0.1,rely=0.8, anchor=W)
    btn2 = Button(translate_toplevel, text="Temizle",font="Times 13 bold",command=temizle)
    btn2.place(relx=0.9,rely=0.8, anchor=E)

    lang_opsiyon = StringVar(translate_toplevel)
    lang_opsiyon.set("\t")

    lang_cekmece = OptionMenu(translate_toplevel,lang_opsiyon,'afrikaans', 'albanian', 'amharic', 'arabic', 'armenian', 'azerbaijani', 'basque', 'belarusian', 'bengali', 'bosnian', 'bulgarian', 'catalan', 'cebuano', 'chichewa', 'chinese (simplified)', 'chinese (traditional)', 'corsican', 'croatian', 'czech', 'danish', 'dutch', 'english', 'esperanto', 'estonian', 'filipino', 'finnish', 'french', 'frisian', 'galician', 'georgian', 'german', 'greek', 'gujarati', 'haitian creole', 'hausa', 'hawaiian', 'hebrew', 'hebrew', 'hindi', 'hmong', 'hungarian', 'icelandic', 'igbo', 'indonesian', 'irish', 'italian', 'japanese', 'javanese', 
    'kannada', 'kazakh', 'khmer', 'korean', 'kurdish (kurmanji)', 'kyrgyz', 'lao', 'latin', 'latvian', 'lithuanian', 'luxembourgish', 'macedonian', 'malagasy', 'malay', 'malayalam', 'maltese', 'maori', 'marathi', 'mongolian', 'myanmar (burmese)', 'nepali', 'norwegian', 'odia', 'pashto', 'persian', 'polish', 'portuguese', 'punjabi', 'romanian', 'russian', 'samoan', 'scots gaelic', 'serbian', 'sesotho', 'shona', 'sindhi', 'sinhala', 'slovak', 'slovenian', 'somali', 'spanish', 'sundanese', 'swahili', 'swedish', 'tajik', 'tamil', 'telugu', 'thai', 'turkish', 'ukrainian', 'urdu', 'uyghur', 'uzbek', 'vietnamese', 'welsh', 'xhosa', 'yiddish', 'yoruba', 'zulu')
    lang_cekmece.place(relx=0.50,rely=0.8, anchor=CENTER)


def encevir():
        global iftrue
        iftrue = False
        en_cumle = english_textbox.get(1.0,END)
        for i in diller:
            if i == lang_opsiyon.get():
                iftrue = True

        if en_cumle == "\n":
            messagebox.showwarning("Hata","Çevrilmesi için soldaki yazı alanına bir değer giriniz...")
        elif (iftrue == False):
            messagebox.showwarning("Hata","Çevrilmesi için dil alanına bir değer giriniz...")
        else:
            output = translator.translate(en_cumle, dest=lang_opsiyon.get())
            turkce_textbox.delete(1.0, END)
            turkce_textbox.insert(1.0, output.text)
            

def temizle():
        english_textbox.delete(1.0,END)
        turkce_textbox.delete(1.0,END)

def yaz(x):
        global yazi
        yazi = yazi + x
        ekran.config(text=yazi)


def hesapla():
    global yazi
    a = eval(yazi)
    ekran.config(text=a)
    yazi = ""


def temizle():
    global yazi
    ekran.config(text="")
    yazi = ""


def sil():
    global yazi
    yazi = yazi[0:-1]
    ekran.config(text=yazi)

def calculator():
    global ekran
    global yazi
    calculator_toplevel = Toplevel()
    calculator_toplevel.geometry("340x280")
    calculator_toplevel.resizable(width=FALSE, height=FALSE)
    calculator_toplevel.title("Hesap Makinesi")

    cerceveana1 = Frame(calculator_toplevel)
    cerceveana1.pack(expand=YES, fill=X)

    cerceveana11 = Frame(cerceveana1)
    cerceveana11.pack(side=TOP, expand=YES, fill=X)

    cerceveana2 = Frame(calculator_toplevel)
    cerceveana2.pack()

    cerceveana21 = Frame(cerceveana2)
    cerceveana21.grid(row=0, column=0)

    cerceveana22 = Frame(cerceveana2)
    cerceveana22.grid(row=0, column=1)

    cerceveana23 = Frame(cerceveana2)
    cerceveana23.grid(row=1, column=0)

    cerceveana24 = Frame(cerceveana2)
    cerceveana24.grid(row=1, column=1)

    cerceve4 = Frame(cerceveana21)
    cerceve4.pack(padx=12)

    cerceve2 = Frame(cerceveana22)
    cerceve2.pack(padx=12)

    yazi = ""
    ekran = Label(cerceveana11)
    ekran.config(textvariable=yazi, relief=SUNKEN, bg="white", height=2, anchor=E)
    ekran.pack(expand=YES, fill=X, padx=12, pady=10)

    c = 0
    d = 0
    e = 0
    for i in (1, 2, 3, 4, 5, 6, 7, 8, 9):
        k = i % 3
        if k == 1:
            Button(cerceve4, text=i, fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
                command=(lambda i=i: yaz(str(i)))).grid(row=c, column=1)
            c = c + 1
        if k == 2:
            Button(cerceve4, text=i, fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
                command=(lambda i=i: yaz(str(i)))).grid(row=d, column=2)
            d = d + 1
        if k == 0:
            Button(cerceve4, text=i, fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
                command=(lambda i=i: yaz(str(i)))).grid(row=e, column=3)
            e = e + 1

    Button(cerceve2, text="X", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("*"))).grid(row=0, column=4)
    Button(cerceve2, text="/", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("/"))).grid(row=1, column=4)
    Button(cerceve2, text="=", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6, command=hesapla).grid(
        row=3, column=4)
    Button(cerceve4, text="(", fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("("))).grid(row=3, column=1)
    Button(cerceve4, text="0", fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("0"))).grid(row=3, column=2)
    Button(cerceve4, text=")", fg="white", bg="black", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz(")"))).grid(row=3, column=3)
    Button(cerceve2, text="AC", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6, command=temizle).grid(
        row=2, column=4)
    Button(cerceve2, text="+", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("+"))).grid(row=0, column=5)
    Button(cerceve2, text="-", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("-"))).grid(row=1, column=5)
    Button(cerceve2, text=".", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6,
        command=(lambda i=i: yaz("."))).grid(row=2, column=5)
    Button(cerceve2, text="C", fg="yellow", bg="Green", font=("Arial", 10, "bold"), height=2, width=6, command=sil).grid(
        row=3, column=5)

def espriypa():
    rastgele_espri = random.choice(espriler)
    if messagebox.askyesno("Espri",message=rastgele_espri):
        espriypa()

def twitter():
    wb.open("https://twitter.com/")

def youtube():
    wb.open("https://youtube.com/")

def instagram():
    wb.open("https://www.instagram.com")

def cmd():
    os.system("start cmd")

def ayarlar():
    os.system("start ms-settings:")

def dosya_gezgini():
    os.system("start C:\ ")

def recyclebin():
    os.system("start shell:RecycleBinFolder")

def facebook():
    wb.open("https://tr-tr.facebook.com")

def whatsapp():
    wb.open("https://web.whatsapp.com/")

def telegram():
    wb.open("https://web.telegram.org/z/")

def discord():
    wb.open("https://https://discord.com/app/")

def twitch():
    wb.open("https://www.twitch.tv")


def create_window_media():
    media_toplevel = Toplevel(bg="#d9d9d9")
    media_toplevel.geometry("350x250")
    media_toplevel.title("Joseph/Sosyal media")
    media_toplevel.minsize(350,250)
    Button(media_toplevel,text="Twitter", command=twitter).place(anchor=CENTER, relx=0.2, rely=0.1, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Youtube", command=youtube).place(anchor=CENTER, relx=0.2, rely=0.22, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Instagram", command=instagram).place(anchor=CENTER, relx=0.2, rely=0.34, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Facebook", command=facebook).place(anchor=CENTER, relx=0.2, rely=0.46, relwidth=0.3, relheight=0.09)
    
    Button(media_toplevel,text="Whatsapp Web", command=whatsapp).place(anchor=CENTER, relx=0.8, rely=0.1, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Telegram", command=telegram).place(anchor=CENTER, relx=0.8, rely=0.22, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Discord", command=discord).place(anchor=CENTER, relx=0.8, rely=0.34, relwidth=0.3, relheight=0.09)
    Button(media_toplevel,text="Twitch", command=twitch).place(anchor=CENTER, relx=0.8, rely=0.46, relwidth=0.3, relheight=0.09)

def create_window_apps():
    apps_toplevel = Toplevel(bg="#d9d9d9")
    apps_toplevel.geometry("350x250")
    apps_toplevel.title("Joseph/Uygulamalar")
    apps_toplevel.minsize(350,250)
    def app_get():
        app_entry_get = app_entry.get()
        command = "start "+app_entry_get
        os.system(command)
    Button(apps_toplevel,text="Komut İstemi", command=cmd).place(anchor=CENTER, relx=0.5, rely=0.1, relwidth=0.4, relheight=0.09)
    Button(apps_toplevel,text="Geri Dönüşüm Kutusu", command=recyclebin).place(anchor=CENTER, relx=0.5, rely=0.22, relwidth=0.4, relheight=0.09)
    Button(apps_toplevel,text="Dosya Gezgini", command=dosya_gezgini).place(anchor=CENTER, relx=0.5, rely=0.34, relwidth=0.4, relheight=0.09)
    Button(apps_toplevel,text="Ayarlar", command=ayarlar).place(anchor=CENTER, relx=0.5, rely=0.46, relwidth=0.4, relheight=0.09)
    app_entry = Entry(apps_toplevel, bg="gray", border=2)
    app_entry.place(relx=0.5,anchor=CENTER,rely=0.65,relheight=0.1,relwidth=0.8)
    Button(apps_toplevel,text="Çalıştır", command=app_get).place(anchor=CENTER, relx=0.5, rely=0.8, relwidth=0.4, relheight=0.09)

def startlist():
    lists_toplevel = Toplevel()
    lists_toplevel.geometry("400x400")
    lists_toplevel.minsize(400,400)
    def ekle():
        eklenecek_var = eklenecek.get()
        if eklenecek_var != "":
            liste.insert(END,eklenecek_var)
            eklenecek.delete(0,END)
    def sil():
        liste.delete(END)
        a = liste.size()
        sheet.delete_rows(a)
    def temizle():
        liste.delete(0,END)
        a = sheet.max_row
        for i in range(a):
            sheet.delete_rows(i)
    
    liste=Listbox(lists_toplevel)
    liste.place(relx=0.5,rely=0.3,relwidth=0.9,relheight=0.5,anchor=CENTER)

    a = sheet.max_row
    a2 = 0

    for i in range(a):
        i = i+1
        liste.insert(END,sheet["A"+str(i)].value)

    def onclosing():
        index = liste.get(0,END)
        index = list(index)
        for i in range(liste.size()):
            i = i+1
            konum = "A{}".format(i)
            if str(index[i-1]) != "":
                sheet[konum] = str(index[i-1])

        if os.path.exists("Liste.xlsx"):
            os.remove("Liste.xlsx")
        workbook.save("Liste.xlsx")
        lists_toplevel.destroy()

    labelm=Label(lists_toplevel,text="Bir veri giriniz:")
    labelm.place(relx=0.5, rely=0.7,anchor=CENTER)
    eklenecek=Entry(lists_toplevel)
    eklenecek.place(relx=0.5, rely=0.8,anchor=CENTER)
    
    ekle_b=Button(lists_toplevel,text="EKLE",command=ekle)
    ekle_b.place(relx=0.5, rely=0.9,anchor=CENTER)

    sil_b=Button(lists_toplevel,text="SİL",command=sil)
    sil_b.place(relx=0.1, rely=0.9,anchor=CENTER)

    clear_b=Button(lists_toplevel,text="TEMİZLE",command=temizle)
    clear_b.place(relx=0.9, rely=0.9,anchor=CENTER)
    

    lists_toplevel.protocol("WM_DELETE_WINDOW", onclosing)